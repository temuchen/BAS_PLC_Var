from xml.etree import ElementTree as ET
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill
import sys
import time
import os
import math#施工math.ceil()向上取整

# 采用openpyxl
# 行和列计数都是从1开始
# 颜色代码：洋红：ff00ff；大红：ff0000；亮黄：ffff00；灰色：aaaaaa；亮绿：00FF00；亮青：00FFFF；白色：FFFFFF
# 定义写入单元格函数
def write_cell(sheet, row_num, column_num, value):
    sheet.cell(row=row_num, column=column_num).value = value


# 定义读取单元格函数，返回值只有空字符（'')和非空字符之分
def read_cell(sheet, row_num, column_num):
    if sheet.cell(row=row_num, column=column_num).value is None:#空值需要用is None来判断
        return ''
    else:
        return str(sheet.cell(row=row_num, column=column_num).value)


# 单元格填充颜色函数
def fill_cell_color(sheet, row_num, column_num, color):
    fill_patt = PatternFill(fill_type='solid', fgColor=color)
    sheet.cell(row=row_num, column=column_num).fill = fill_patt


# 替换特殊字符
def replace_xchar(source_str):
    result_str = ''
    xchar_str = r'&%#！*—-/?|{}~`^@.""''“”；：:\\ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩⅪⅫ\[\]'
    for x_char in source_str:
        if x_char in xchar_str:
            source_str = source_str.replace(x_char, '_')
        if x_char in ['(', ')', '（', '）',' ']:
            source_str = source_str.replace(x_char, '')
    result_str = source_str
    return result_str
#删除表格内容
def del_sht_content(sheet, row_num, col_num):
    for i in range(2, row_num+1):
        for j in range(2, col_num+1):
            sheet.cell(i,j).value = ''

# Excel查重处理
def find_duplication(sheet, column_num):
    ls_dup = []
    for i in range(2, sheet.max_row):
        if read_cell(sheet, i, column_num) !='' and (
                read_cell(sheet, i, column_num) == read_cell(sheet, i + 1, column_num) or \
                read_cell(sheet, i, column_num + 1) == read_cell(sheet, i + 1, column_num + 1)):
            ls_dup.append('第{0}行重复'.format(i))
    return ls_dup
#将一个sheet的内容复制到另外一个sheet中
def copy_sheet_content(sht_source, sht_dest, columen_count):
    for i in range(sht_source.max_row):
        for j in range(columen_count):
            sht_dest.cell(row=i+1, column=j+1).value = sht_source.cell(row=i+1, column=j+1).value
#读取设备实例表，生成设备list和设备dic
def read_dev_sht(sht_dev, ls_kind):
    temp_ls_dev = []
    temp_dic_dev = {}
    for i in range(2, sht_dev.max_row + 1):
        fill_cell_color(sht_dev, i, 2, 'FFFFFF')  # 表格填充白色
        if read_cell(sht_dev, i, 2) in ls_kind:
            temp_ls_dev.append(read_cell(sht_dev, i, 4))
            temp_dic_dev[read_cell(sht_dev, i, 4)] = [read_cell(sht_dev, i, 2),  # 0:设备类型
                                                 read_cell(sht_dev, i, 3),  # 1:设备名称
                                                 read_cell(sht_dev, i, 6),  # 2:设备安装位置
                                                 read_cell(sht_dev, i, 7),  # 3:所属系统
                                                 read_cell(sht_dev, i, 8),  # 4:从属区域
                                                 read_cell(sht_dev, i, 9),  # 5:对应的功能块
                                                 read_cell(sht_dev, i, 10),  # 6:模式工艺分区索引
                                                 read_cell(sht_dev, i, 11), # 7:连锁设备
                                                 read_cell(sht_dev, i, 5)]  # 8:设备编号
            fill_cell_color(sht_dev, i, 2, '00FFFF')  # 读取到的设备填充亮青色
    return temp_ls_dev, temp_dic_dev

# s:原字符串  length：填充总长度
#左对齐
def standard_string(s,length):
    Count=0
    for aim in s:
        if('\u4e00' <= aim <= '\u9fff'):
            Count+=1
    flag=length-len(s)-Count
    return s+' '*flag

current_path = os.getcwd()
excel_file_path = input("请输入BAS实例表文件路径，仅支持xlsx格式，例如：d:\\test.xlsx：")
if excel_file_path == '':
    excel_file_path = 'd:\\test.xlsx'


# 获取工作簿对象
wb_mode = openpyxl.load_workbook(excel_file_path)
# 获取所有工作表
ls_shenames = wb_mode.sheetnames
# print(ls_shenames)
sht_zone = wb_mode['车站工艺分区清单']

#自动计算各个工艺分区的模式数量和模式动作序列长度
#可以直接读取模式动作序列表格的第一个sheet页对应的单元格
# row_index = 1
# column_index = 1
# for i in range(sht_zone.max_row):
#     temp_zone_name = read_cell(sht_zone,i+2,2)#工艺分区名称
#     if temp_zone_name != '':
#         act_len = int(wb_mode[temp_zone_name].max_column/4)+2#该长度包含了模式号的长度
#         mode_quan = wb_mode[temp_zone_name].max_row+3#模式数量
#         write_cell(sht_zone,row_index+1,3,act_len)
#         write_cell(sht_zone,row_index+1,4,mode_quan)
#         row_index = row_index + 1


#读取各个工艺分区的模式长度、数量、关键字
dic_modeinfo = {}
ls_zonename = []
for i in range(sht_zone.max_row):
    temp_zone_name = read_cell(sht_zone,i+2,2)
    if temp_zone_name != '':
        act_len = int(read_cell(sht_zone,i+2,3))
        mode_quan = int(read_cell(sht_zone,i+2,4))
        zone_sn = read_cell(sht_zone,i+2,5)
        dev_quan = int(read_cell(sht_zone,i+2,6))
        dic_modeinfo[temp_zone_name]=[act_len,mode_quan,zone_sn,dev_quan]#序列长度、模式数量、分区前缀、设备数量
        ls_zonename.append(temp_zone_name)

print(dic_modeinfo)

# input()
# exit(0)
# -------------------生成变量---------------------------
# ------------------------------------------------
# 直接生成变量导入文件
# 2019-7-31，张海波
# -------------------------------------------------
# #-----------------------------------------------
# 创建节点：root = ET.Element('Root')
# 创建文档：tree = ET.ElementTree(root)
# 设置文本值：element.text = 'default'
# 设置属性：element.set('age', str(i))
# 添加节点：root.append(element)
# 写入文档：tree.write('default.xml', encoding='utf-8', xml_declaration=True)
# nodes = node[1:5]
# node.append(subnode)
# node.insert(0, subnode)
# node.remove(subnode)
# ---------------------
# 作者：新安浅滩
# 来源：CSDN
# 原文：https://blog.csdn.net/hu694028833/article/details/81089959
# 版权声明：本文为博主原创文章，转载请附上博文链接！
# -------------------------------------------------------------------
# 增加换行符，保持队形


def __indent(elem, level=0):
    i = "\n" + level * "\t"
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "\t"
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            __indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i

# 创建具体的变量实例
def create_var_node(name, typeName, topologicalAddress, comment,namestr,namecomment):
    instanceED_node = ET.Element('instanceElementDesc',{'name':namestr})
    comment_node = ET.Element('comment')
    comment_node.text = comment
    if namecomment != '':
        insEDComment_node = ET.Element('comment')
        insEDComment_node.text = namecomment
        instanceED_node.append(insEDComment_node)
    if topologicalAddress != '':
        dev_node = ET.Element('variables',
                              {'name': name, 'typeName': typeName, 'topologicalAddress': topologicalAddress})
        dev_node.append(comment_node)
        if namestr != '':
            dev_node.append(instanceED_node)
    else:
        dev_node = ET.Element('variables', {'name': name, 'typeName': typeName})
        dev_node.append(comment_node)
    return dev_node

#创建数组元素注释
''' 
<instanceElementDesc name="[0]">
    <comment>ceshi</comment>
</instanceElementDesc>
'''
def create_desc_node(namestr, comment):
    instanceED_node = ET.Element('instanceElementDesc',{'name':namestr})
    comment_node = ET.Element('comment')
    comment_node.text = comment
    instanceED_node.append(comment_node)
    return instanceED_node


def create_xml_head():
    # 创建根节点
    root = ET.Element('VariablesExchangeFile')
    # 创建变量导入文件头
    file_header01 = ET.Element('fileHeader', {'company': 'Schneider Automation',
                                              'product': 'Unity Pro XL V11.1 - 160831F',
                                              'dateTime': 'date_and_time#2017-5-27-17:25:52',
                                              'content': '变量源文件',
                                              'DTDVersion': '41'})
    file_header02 = ET.Element('contentHeader', {'name': '项目',
                                                 'version': '0.0.101',
                                                 'dateTime': 'date_and_time#2017-5-19-13:53:47'})
    root.append(file_header01)
    root.append(file_header02)

    return root

data_block = ET.Element('dataBlock')

temp_index = 0
ls_zoneinfostr = ['火灾模式标志','阻塞模式标志','当前执行模式号','灾后恢复时执行的模式号',
                  '模式命令来源：1-IBP；2-ISCS;3-FAS;4-时间表;5-联动；6-焓值；7-灾后恢复',
                  '模式执行状态','模式控制方式','模式执行成功','模式执行失败','模式正在执行']
modeinfoaddress = 6000
modecheckadd = modeinfoaddress + 1000
mode_add_index = 0
for zn in ls_zonename:
    #给综合监控的模式信息
    varname = dic_modeinfo[zn][2]+'_ModeInfo'
    typename = 'udt_ModeInfo2ISCS'
    topologicalAddress = '%MW' + str(modeinfoaddress)#导入程序后需要手动修改
    comment = zn+'模式信息给ISCS'
    namestr = ''
    namecomment = ''
    modeinfo2iscs_node = create_var_node(varname, typename, topologicalAddress,comment,namestr,namecomment)
    data_block.append(modeinfo2iscs_node)
    modeinfoaddress = modeinfoaddress + 9

for zn in ls_zonename:
    #模式查看变量
    varname = dic_modeinfo[zn][2]+'_ModeCheckISCS'
    typename = 'array[0..1] of int'
    topologicalAddress = '%MW' + str(modecheckadd)#导入程序后需要手动修改
    comment = zn+'模式动作内容查看'
    namestr = ''
    namecomment = ''
    modecheck_node = create_var_node(varname, typename, topologicalAddress,comment,namestr,namecomment)
    data_block.append(modecheck_node)
    modecheckadd = modecheckadd + 2


#创建模式动作序列变量
for zn in ls_zonename:
    varname = dic_modeinfo[zn][2]+'_ModeActSeq'
    typename = 'array[0..'+str(dic_modeinfo[zn][0]*dic_modeinfo[zn][1]-1) + '] of int'#序列长度*模式数量-1
    topologicalAddress = ''
    comment = zn+'动作序列'
    namestr = ''
    namecomment = ''
    modeactseq_node = create_var_node(varname, typename, topologicalAddress,comment,namestr,namecomment)
    temp_index = 0
    for i in range(dic_modeinfo[zn][1]):
        for j in range(dic_modeinfo[zn][0]):
            desc_name = '[' + str(temp_index) + ']'
            if j == 0:
                desc_comment = '第'+str(i+1)+'个模式号'
            else:
                desc_comment = str(j)
            instanceElementDesc_node = create_desc_node(desc_name, desc_comment)
            modeactseq_node.append(instanceElementDesc_node)
            temp_index = temp_index + 1
    data_block.append(modeactseq_node)


    # 模式不符信息
    varname = dic_modeinfo[zn][2]+'_ModeRefuseInfo'
    typename = 'array[0..'+str(math.ceil(dic_modeinfo[zn][3]/16)+1) + '] of int'#设备数量/16+1
    topologicalAddress = ''
    comment = zn+'模式不符信息'
    namestr = ''
    namecomment = ''
    moderefuseinfo_node = create_var_node(varname, typename, topologicalAddress,comment,namestr,namecomment)
    data_block.append(moderefuseinfo_node)

    # 模式控制命令
    varname = dic_modeinfo[zn][2] + '_ModeCmd'
    typename = 'array[0..99,0..3] OF BOOL'
    topologicalAddress = ''
    comment = zn + '模式控制命令'
    namestr = ''
    namecomment = ''
    modecmd_node = create_var_node(varname, typename, topologicalAddress, comment, namestr, namecomment)
    data_block.append(modecmd_node)

    # 模式执行状态，仅程序内部使用
    varname = dic_modeinfo[zn][2] + '_ExcuteSt'
    typename = 'array[0..15] OF BOOL'
    topologicalAddress = ''
    comment = zn + '模式执行状态bit0：正在执行；bit1：执行失败；bit2：执行成功'
    namestr = ''
    namecomment = ''
    modeexcute_node = create_var_node(varname, typename, topologicalAddress, comment, namestr, namecomment)
    data_block.append(modeexcute_node)

    # 模式控制命令查看
    varname = dic_modeinfo[zn][2] + '_ModeCmdQuery'
    typename = 'array[0..99,0..3] OF BOOL'
    topologicalAddress = ''
    comment = zn + '模式表动作查询'
    namestr = ''
    namecomment = ''
    modecmdquery_node = create_var_node(varname, typename, topologicalAddress, comment, namestr, namecomment)
    data_block.append(modecmdquery_node)

    # 模式命令源变量，每个工艺分区创建一个
    varname = dic_modeinfo[zn][2] + '_ModeCmdSource'
    typename = 'udt_ModeCmdSource'
    comment = zn + '模式命令源'
    modecmdsource_node = create_var_node(varname, typename, '',comment, '', '')
    data_block.append(modecmdsource_node)

    #模式控制功能块变量——ModeASSearch
    varname = 'ModeASSearch_'+dic_modeinfo[zn][2]
    typename = 'aoi_ModeASSearch_New'
    comment = zn + '模式搜索'
    namestr = ''
    namecomment = ''
    modeassearch_fbd_node = create_var_node(varname, typename, topologicalAddress, comment, namestr, namecomment)
    data_block.append(modeassearch_fbd_node)

    #模式控制功能块变量——ModeCalculateLogic
    varname = 'ModeCalculateLogic_'+dic_modeinfo[zn][2]
    typename = 'aoi_ModeCalculateLogic_New'
    comment = zn + '模式计算'
    namestr = ''
    namecomment = ''
    ModeCalculateLogic_fbd_node = create_var_node(varname, typename, topologicalAddress, comment, namestr, namecomment)
    data_block.append(ModeCalculateLogic_fbd_node)

    #模式控制功能块变量——ModeExcute
    varname = 'ModeExcute_'+dic_modeinfo[zn][2]
    typename = 'aoi_ModeExcute_New'
    comment = zn + '模式执行'
    namestr = ''
    namecomment = ''
    ModeExcute_fbd_node = create_var_node(varname, typename, topologicalAddress, comment, namestr, namecomment)
    data_block.append(ModeExcute_fbd_node)

# IBP使能信号
varname = 'bIBPEnable'
typename = 'bool'
comment = 'IBP使能'
namestr = ''
namecomment = ''
IBP_fbd_node = create_var_node(varname, typename, topologicalAddress, comment, namestr, namecomment)
data_block.append(IBP_fbd_node)

#计算出的模式号变量
temp_index = 0
modecal_node = create_var_node('ModeCalculationCode','array[0..29] of int','','计算出的模式号','','')
for zn in ls_zonename:
        desc_name = '[' + str(temp_index) + ']'
        desc_comment = zn
        instanceElementDesc_node = create_desc_node(desc_name,desc_comment)
        modecal_node.append(instanceElementDesc_node)
        temp_index = temp_index + 1
data_block.append(modecal_node)

# 模式控制功能块变量——ModeInfoRefresh
varname = 'ModeInfoRefresh'
typename = 'aoi_ModeInfoRefresh_New'
comment = '模式信息更新'
namestr = ''
namecomment = ''
ModeInfoRefresh_fbd_node = create_var_node(varname, typename, topologicalAddress, comment, namestr, namecomment)
data_block.append(ModeInfoRefresh_fbd_node)

# 模式控制功能块变量——aoi_ModeLinkage
varname = 'ModeLinkage'
typename = 'aoi_ModeLinkage_New'
comment = '模式联动'
namestr = ''
namecomment = ''
ModeLinkage_fbd_node = create_var_node(varname, typename, topologicalAddress, comment, namestr, namecomment)
data_block.append(ModeLinkage_fbd_node)

# 模式控制功能块变量——ModeRecover
varname = 'ModeRecover'
typename = 'aoi_ModeRecover_New'
comment = '模式复位'
namestr = ''
namecomment = ''
ModeRecover_fbd_node = create_var_node(varname, typename, topologicalAddress, comment, namestr, namecomment)
data_block.append(ModeRecover_fbd_node)



#模式联动规则
temp_index = 0
typename = 'array[0..' + str(len(ls_zonename) * 100 - 1) + '] of int'  # 工艺分区数量*100-1，最多100条联动规则
modelinkage_node = create_var_node('ModeLinkageRule',typename,'','模式联动规则','','')
for i in range(len(ls_zonename) * 100):
    desc_name = '[' + str(temp_index) + ']'
    desc_comment = ls_zonename[i%len(ls_zonename)]
    instanceElementDesc_node = create_desc_node(desc_name,desc_comment)
    modelinkage_node.append(instanceElementDesc_node)
    temp_index = temp_index + 1
data_block.append(modelinkage_node)

#工艺分区模式信息
temp_index = 0
typename = 'array[0..' + str(len(ls_zonename) * 10 - 1) + '] of int'  # 工艺分区数量*10-1
modezoneinfo_node = create_var_node('ModeZoneInfo',typename,'','工艺分区模式信息','','')
for zn in ls_zonename:
    for txt in ls_zoneinfostr:
        desc_name = '[' + str(temp_index) + ']'
        desc_comment = zn+txt
        instanceElementDesc_node = create_desc_node(desc_name,desc_comment)
        modezoneinfo_node.append(instanceElementDesc_node)
        temp_index = temp_index + 1

data_block.append(modezoneinfo_node)


mode_root = create_xml_head()
mode_root.append(data_block)

__indent(mode_root)
tree = ET.ElementTree(mode_root)
tree.write(current_path+'\\var_mode.xsy', encoding='utf-8', short_empty_elements=False, xml_declaration=True)



print('程序变量生成完毕，存储地址为当前工具所在目录下的“var_mode.xsy”。')
print('继续生成程序段请按3，其他键退出：')
input()

# ------------------------生成程序代码------------------------------------------------------------------------------
# print(opt_select)
# if opt_select == '3':
#     ls_dev, dic_dev = read_dev_sht(sht_dev)
# print(ls_dev)
# print(dic_dev)

def creat_xml_node(root_name, dic_node_content):  # dic_node_content是字典类型
    if len(dic_node_content) == 0:
        return ET.Element(root_name)
    else:
        return ET.Element(root_name, dic_node_content)


# 创建根节点，生成以下代码：
# <FBDExchangeFile>
# 	<fileHeader company="Schneider Automation" product="Unity Pro XL V11.1 - 160831F" dateTime="date_and_time#2019-8-23-16:25:20" content="导出的功能块源文件" DTDVersion="41"></fileHeader>
# 	<contentHeader name="项目" version="0.0.282" dateTime="date_and_time#2018-3-10-10:31:8"></contentHeader>
# 	<program>
# 		<identProgram name="DevControl_SLF" type="section" task="MAST"></identProgram>
# 		<FBDSource nbRows="24" nbColumns="36">
# 			<networkFBD>


def create_prg_head_node(prg_name):
    prg_root = creat_xml_node('FBDExchangeFile', {})
    dic_fileHeader_content = {'company': 'Schneider Automation', 'product': 'Unity Pro XL V11.1 - 160831F',
                              'dateTime': 'date_and_time#2019-8-20-8:5:2',
                              'content': '导出的功能块源文件', 'DTDVersion': '41'}
    dic_contentHeader_content = {'name': '项目', 'version': '0.0.300', 'dateTime': 'date_and_time#2019-7-19-13:22:18'}

    prg_file_header = creat_xml_node('fileHeader', dic_fileHeader_content)
    prg_content_header = creat_xml_node('contentHeader', dic_contentHeader_content)
    prg_program = creat_xml_node('program', {})
    prg_indentProgram = creat_xml_node('identProgram', {'name':prg_name, 'type': 'section', 'task': 'MAST'})
    prg_FBDSource = creat_xml_node('FBDSource', {'nbRows': '24', 'nbColumns': '36'})
    prg_networkFBD = creat_xml_node('networkFBD', {})
    #合并结构
    prg_FBDSource.append(prg_networkFBD)
    prg_program.append(prg_indentProgram)
    prg_program.append(prg_FBDSource)
    prg_root.append(prg_file_header)
    prg_root.append(prg_content_header)
    prg_root.append(prg_program)

    return  prg_root

#生成以下代码
#<FFBBlock instanceName="SL_01_01" typeName="Dev_SLF" additionnalPinNumber="0" enEnO="false" width="15" height="16">
#宽度可以调节，高度根据功能块默认。
def creat_FFBBlock_node(instanceName, typeName, width, height, posX, posY):
    prg_FBDBlock = creat_xml_node('FFBBlock', {'instanceName': instanceName,
                                               'typeName': typeName, 'additionnalPinNumber': '0',
                                               'enEnO': 'false', 'width': width, 'height': height})
    prg_objPosition = creat_xml_node('objPosition', {'posX': posX, 'posY': posY})
    prg_FBDBlock.append(prg_objPosition)
    return prg_FBDBlock

# 创建设备块引脚信息节点

def create_ffb_pin(pin_kind, formal_parameter, effective_parameter):
    if effective_parameter == '':
        if pin_kind == 1:#输入参数
            return creat_xml_node('inputVariable', {'invertedPin': 'false', 'formalParameter': formal_parameter})
        elif pin_kind == 2:#输出参数
            return creat_xml_node('outputVariable', {'invertedPin': 'false', 'formalParameter': formal_parameter})
        else:
            print('引脚参数设置错误，请检查,按任意键退出。')
            input()
            exit()
    else:
        if pin_kind == 1:
            return creat_xml_node('inputVariable', {'invertedPin': 'false', 'formalParameter': formal_parameter,
                                                    'effectiveParameter': effective_parameter})
        elif pin_kind == 2:
            return creat_xml_node('outputVariable', {'invertedPin': 'false', 'formalParameter': formal_parameter,
                                                     'effectiveParameter': effective_parameter})
        else:
            print('引脚参数设置错误，请检查,按任意键退出。')
            input()
            exit()

# 各种设备类型功能块定义
#模式计算功能
def create_modecal_node(stationinfo, zoneinfo, mode_index):
    prg_descriptionFFB = creat_xml_node('descriptionFFB', {'execAfter': ''})
    prg_descriptionFFB.append( create_ffb_pin(1, 'EN', ''))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'bStationFireRecoverFlag', stationinfo + '[2]'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'bStationFireFlag', stationinfo + '[0]'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'bStationBlockFlag', stationinfo + '[1]'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'iCtrMode', zoneinfo+'['+str(mode_index*10 + 6)+']'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'CmdSource', zoneinfo+'['+str(mode_index*10 + 4)+']'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'iCurrentModeCode', zoneinfo+'['+str(mode_index*10 + 2)+']'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ModeSource', 'ModeCmdSource'+'['+str(mode_index)+']'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'iModeCalculated', 'ModeCalculationCode'+'['+str(mode_index)+']'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'bIBPEnable', 'bIBPEnable'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(2, 'ENO', ''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'iCtrMode',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'CmdSource',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'iCurrentModeCode',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'ModeSource',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'iModeCalculated',''))

    return prg_descriptionFFB

#模式联动功能块
def create_modelinkage_node(ZoneQuantity):
    prg_descriptionFFB = creat_xml_node('descriptionFFB', {'execAfter': ''})
    prg_descriptionFFB.append( create_ffb_pin(1, 'EN', ''))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'CalCode', 'ModeCalculationCode'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'MatrixRule', 'ModeLinkageRule'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'StationModeInfo', 'ModeStationInfo'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ZoneModeInfo', 'ModeZoneInfo'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ZoneQuantity', str(ZoneQuantity)))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(2, 'ENO', ''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'CalCode',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'MatrixRule',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'StationModeInfo',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'ZoneModeInfo',''))

    return prg_descriptionFFB

#模式灾后恢复模块
def create_moderecover_node(ZoneQuantity):
    prg_descriptionFFB = creat_xml_node('descriptionFFB', {'execAfter': ''})
    prg_descriptionFFB.append( create_ffb_pin(1, 'EN', ''))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'StationModeInfo', 'ModeStationInfo'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'CalCode', 'ModeCalculationCode'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'FireSignal', 'ModeFireSignal'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ZoneModeInfo', 'ModeZoneInfo'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ZoneQuantity', str(ZoneQuantity)))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(2, 'ENO', ''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'StationModeInfo',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'CalCode',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'ZoneModeInfo',''))

    return prg_descriptionFFB

#模式执行模块
def create_modeexcute_node(zonename, mode_index):
    prg_descriptionFFB = creat_xml_node('descriptionFFB', {'execAfter': ''})
    prg_descriptionFFB.append( create_ffb_pin(1, 'EN', ''))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'tModeExcuteTime', 't#60s'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'RefuseInfo', zonename + '_ModeRefuseInfo'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'CalCode', 'ModeCalculationCode[' + str(mode_index) +']'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'CurrentCode', 'ModeZoneInfo[' + str(mode_index*10 + 2) +']'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ActSta', 'ModeZoneInfo[' + str(mode_index*10 + 2) + ']'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(2, 'ENO', ''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'CalCode',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'CurrentCode',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'ActSta',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'ExcuteSt',zonename + '_ExcuteSt'))

    return prg_descriptionFFB

#模式搜索功能块
def create_modeseach_node(zonename, mode_index, modequant, aslen):
    prg_descriptionFFB = creat_xml_node('descriptionFFB', {'execAfter': ''})
    prg_descriptionFFB.append( create_ffb_pin(1, 'EN', ''))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'bModeExcuteSignal', zonename + '_ExcuteSt[0]'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'iModeCode',  'ModeZoneInfo[' + str(mode_index*10 + 2) +']'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ModeASArray', zonename + '_ModeActSeq'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ModeQuantity', str(modequant)))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ASLen', str(aslen)))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(2, 'ENO', ''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'ModeCmd', zonename + '_ModeCmd'))
    prg_descriptionFFB.append(create_ffb_pin(2, 'bNotFindFlag',''))

    return prg_descriptionFFB

#模式刷新功能块
def create_moderefresh_node(ZoneQuantity):
    prg_descriptionFFB = creat_xml_node('descriptionFFB', {'execAfter': ''})
    prg_descriptionFFB.append( create_ffb_pin(1, 'EN', ''))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'StationModeInfo', 'ModeStationInfo'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ZoneQuantity', str(ZoneQuantity)))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ZoneModeInfo', 'ModeZoneInfo'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(2, 'ENO', ''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'StationModeInfo',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'ZoneModeInfo',''))

    return prg_descriptionFFB


#模式查看功能块
def create_modelookup_node(zonename, modequant, aslen):
    prg_descriptionFFB = creat_xml_node('descriptionFFB', {'execAfter': ''})
    prg_descriptionFFB.append( create_ffb_pin(1, 'EN', ''))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'bModeExcuteSignal', '1'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'iModeCode',  zonename + '_ModeCheckISCS[0]'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ModeASArray', zonename + '_ModeActSeq'))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ModeQuantity', str(modequant)))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(1, 'ASLen', str(aslen)))#通用引脚
    prg_descriptionFFB.append(create_ffb_pin(2, 'ENO', ''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'ModeCmd', zonename + '_ModeCmdQuery'))
    prg_descriptionFFB.append(create_ffb_pin(2, 'bNotFindFlag',''))
    prg_descriptionFFB.append(create_ffb_pin(2, 'CodeToISCS', zonename + '_ModeCheckISCS[1]'))

    return prg_descriptionFFB

stationinfo = 'ModeStationInfo'
zoneinfo = 'ModeZoneInfo'
modeindex = 0
x_count = 0
y_count = 0

ls_mode_cal = []
x_index_cal = 15
y_index_cal = 5
ls_mode_linkage = []
x_index_linkage = 15
y_index_linkage = 50
ls_mode_recover = []
x_index_recover = 15
y_index_recover = 70
ls_mode_excute = []
x_index_excute = 15
y_index_excute = 90
ls_mode_search = []
x_index_search = 15
y_index_search = 130
ls_mode_refresh = []
x_index_refresh = 15
y_index_refresh = 160
ls_mode_lookup = []
x_index_lookup = 15
y_index_lookup = 190

for zn in ls_zonename:
    #模式计算
    temp_et_node_dev = create_modecal_node(stationinfo, zoneinfo, modeindex)
    temp_et_node_ffb = creat_FFBBlock_node('ModeCalculateLogic_' + dic_modeinfo[zn][2], 'aoi_ModeCalculateLogic_New', str(20), str(13),
                                           str(x_index_cal + (x_count % 10) * 35), str(y_index_cal + (y_count % 15) * 15))
    temp_et_node_ffb.append(temp_et_node_dev)
    ls_mode_cal.append(temp_et_node_ffb)

    #模式执行
    temp_et_node_dev = create_modeexcute_node(dic_modeinfo[zn][2],  modeindex)
    temp_et_node_ffb = creat_FFBBlock_node('ModeExcute_' + dic_modeinfo[zn][2], 'aoi_ModeExcute_New', str(16), str(13),
                                           str(x_index_excute + (x_count % 10) * 35), str(y_index_excute + (y_count % 15) * 15))
    temp_et_node_ffb.append(temp_et_node_dev)
    ls_mode_excute.append(temp_et_node_ffb)


    #模式搜索
    temp_et_node_dev = create_modeseach_node(dic_modeinfo[zn][2], modeindex, dic_modeinfo[zn][1], dic_modeinfo[zn][0])
    temp_et_node_ffb = creat_FFBBlock_node('ModeASSearch_' + dic_modeinfo[zn][2], 'aoi_ModeASSearch_New', str(17), str(9),
                                           str(x_index_search + (x_count % 10) * 35), str(y_index_search + (y_count % 15) * 15))
    temp_et_node_ffb.append(temp_et_node_dev)
    ls_mode_search.append(temp_et_node_ffb)

    #模式查看
    temp_et_node_dev = create_modelookup_node(dic_modeinfo[zn][2], dic_modeinfo[zn][1], dic_modeinfo[zn][0])
    temp_et_node_ffb = creat_FFBBlock_node('ModeLookup_' + dic_modeinfo[zn][2], 'aoi_ModeLookup', str(17), str(9),
                                           str(x_index_lookup + (x_count % 10) * 35), str(y_index_lookup + (y_count % 15) * 15))
    temp_et_node_ffb.append(temp_et_node_dev)
    ls_mode_lookup.append(temp_et_node_ffb)

    x_count = x_count + 1
    y_count = int(x_count/10)

    modeindex = modeindex + 1

#模式联动
temp_et_node_dev = create_modelinkage_node(len(ls_zonename))
temp_et_node_ffb = creat_FFBBlock_node('ModeLinkage', 'aoi_ModeLinkage_New', str(17), str(9),
                                       str(x_index_linkage), str(y_index_linkage))
temp_et_node_ffb.append(temp_et_node_dev)
ls_mode_linkage.append(temp_et_node_ffb)

#模式复位
temp_et_node_dev = create_moderecover_node(len(ls_zonename))
temp_et_node_ffb = creat_FFBBlock_node('ModeRecover', 'aoi_ModeRecover_New', str(17), str(9),
                                       str(x_index_recover), str(y_index_recover))
temp_et_node_ffb.append(temp_et_node_dev)
ls_mode_recover.append(temp_et_node_ffb)

#模式刷新
temp_et_node_dev = create_moderefresh_node(len(ls_zonename))
temp_et_node_ffb = creat_FFBBlock_node('ModeInfoRefresh', 'aoi_ModeInfoRefresh_New', str(20), str(13),
                                       str(x_index_refresh), str(y_index_refresh))
temp_et_node_ffb.append(temp_et_node_dev)
ls_mode_refresh.append(temp_et_node_ffb)



#模式控制程序
prg_root_modectr = create_prg_head_node('ModeCtr')  # 文件根节点
for nb in prg_root_modectr.iter('networkFBD'):
    for nd in ls_mode_cal:
        nb.append(nd)
    for nd in ls_mode_linkage:
        nb.append(nd)
    for nd in ls_mode_recover:
        nb.append(nd)
    for nd in ls_mode_excute:
        nb.append(nd)
    for nd in ls_mode_search:
        nb.append(nd)
    for nd in ls_mode_refresh:
        nb.append(nd)
    for nd in ls_mode_lookup:
        nb.append(nd)

__indent(prg_root_modectr)
prg_tree = ET.ElementTree(prg_root_modectr)  # 转化为树
prg_tree.write(current_path + '\\prg_ModeCtr.xbd', encoding='utf-8', short_empty_elements=False, xml_declaration=True)

print("程序段生成完毕，带有‘AEnd’的请导入主端PLC，‘BEnd’的导入从端PLC！")
print('按任意键退出！')
input()

